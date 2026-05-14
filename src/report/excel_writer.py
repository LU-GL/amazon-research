"""Excel报告生成模块。

用openpyxl生成格式化的竞品矩阵Excel工作簿。
"""
import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="微软雅黑", size=10)
PAIN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PRAISE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def write_competitive_excel(
    matrix_df: pd.DataFrame,
    sentiment_data: dict[str, dict],
    comparison: dict | None = None,
    output_path: Path | None = None,
) -> Path:
    """生成竞品矩阵Excel工作簿。"""
    wb = Workbook()

    # Sheet 1: 竞品矩阵
    ws1 = wb.active
    ws1.title = "竞品矩阵"
    matrix_cols = [
        "ASIN", "品牌", "标题", "价格", "评分", "评论数", "BSR",
        "核心痛点1", "痛点1频次", "核心痛点2", "痛点2频次", "核心痛点3", "痛点3频次",
        "好评主题1", "好评主题2", "功能需求", "质量问题", "整体评价", "关键洞察",
    ]
    _write_dataframe_sheet(ws1, matrix_df, matrix_cols)
    _set_col_widths(ws1, [15, 12, 40, 10, 8, 10, 25, 20, 10, 20, 10, 20, 10, 20, 20, 25, 20, 10, 30])

    # Sheet 2: 痛点分析
    ws2 = wb.create_sheet("痛点分析")
    pain_rows = []
    for asin, analysis in sentiment_data.items():
        for pp in analysis.get("pain_points", []):
            pain_rows.append({
                "ASIN": asin,
                "痛点主题": pp["theme"],
                "影响程度": pp.get("severity", ""),
                "出现频次": pp.get("frequency", 0),
                "原文摘录": " | ".join(pp.get("examples", [])[:3]),
            })
    pain_df = pd.DataFrame(pain_rows)
    if not pain_df.empty:
        pain_df = pain_df.sort_values("出现频次", ascending=False)
    _write_dataframe_sheet(ws2, pain_df, ["ASIN", "痛点主题", "影响程度", "出现频次", "原文摘录"])

    # Sheet 3: 好评分析
    ws3 = wb.create_sheet("好评分析")
    praise_rows = []
    for asin, analysis in sentiment_data.items():
        for pp in analysis.get("praise_points", []):
            praise_rows.append({
                "ASIN": asin,
                "好评主题": pp["theme"],
                "出现频次": pp.get("frequency", 0),
                "原文摘录": " | ".join(pp.get("examples", [])[:3]),
            })
    praise_df = pd.DataFrame(praise_rows)
    if not praise_df.empty:
        praise_df = praise_df.sort_values("出现频次", ascending=False)
    _write_dataframe_sheet(ws3, praise_df, ["ASIN", "好评主题", "出现频次", "原文摘录"])

    # Sheet 4: 机会点
    ws4 = wb.create_sheet("机会点")
    opportunity_rows = []
    for asin, analysis in sentiment_data.items():
        for fr in analysis.get("feature_requests", []):
            opportunity_rows.append({
                "ASIN": asin,
                "类型": "功能需求",
                "主题": fr["theme"],
                "频次": fr.get("frequency", 0),
                "原文摘录": " | ".join(fr.get("examples", [])[:3]),
            })
        for qi in analysis.get("quality_issues", []):
            opportunity_rows.append({
                "ASIN": asin,
                "类型": "质量问题",
                "主题": qi["theme"],
                "频次": qi.get("frequency", 0),
                "严重度": qi.get("severity", ""),
            })
    opp_df = pd.DataFrame(opportunity_rows) if opportunity_rows else pd.DataFrame()
    if not opp_df.empty:
        opp_df = opp_df.sort_values("频次", ascending=False)
    _write_dataframe_sheet(ws4, opp_df, list(opp_df.columns) if not opp_df.empty else [])

    # Sheet 5: 原始数据
    ws5 = wb.create_sheet("原始数据")
    ws5.cell(row=1, column=1, value="情感分析原始JSON数据")
    ws5.cell(row=1, column=1).font = Font(bold=True, size=12)
    row_num = 3
    for asin, analysis in sentiment_data.items():
        ws5.cell(row=row_num, column=1, value=f"ASIN: {asin}")
        ws5.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        json_text = json.dumps(analysis, ensure_ascii=False, indent=2)
        for line in json_text.split("\n"):
            ws5.cell(row=row_num, column=1, value=line)
            row_num += 1
        row_num += 1

    # 保存
    if output_path is None:
        from src.config import OUTPUT_DIR
        output_path = OUTPUT_DIR / "competitive_matrix.xlsx"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Excel已保存: {output_path}")
    return output_path


def _write_dataframe_sheet(ws, df: pd.DataFrame, columns: list[str]):
    """将DataFrame写入工作表并添加格式。"""
    if df.empty:
        ws.cell(row=1, column=1, value="暂无数据")
        return

    # 只保留需要的列（如果存在）
    available_cols = [c for c in columns if c in df.columns]
    if available_cols:
        df_out = df[available_cols].copy()
    else:
        df_out = df.copy()

    # 写表头
    for col_idx, col_name in enumerate(df_out.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 写数据
    for row_idx, (_, row) in enumerate(df_out.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = ws.dimensions


def _set_col_widths(ws, widths: list[int]):
    """设置列宽。"""
    for i, width in enumerate(widths, 1):
        if i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = width
